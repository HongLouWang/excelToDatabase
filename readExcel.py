#############################################################################################
#Author:    Shutian Wang
#Date:      Nov. 2 2021
#Desc:      Load Excel and Pass it to UI part
#Project:   Load Excel To Database
#############################################################################################

from os.path import exists
from openpyxl import load_workbook

class excel():

    workbook = ""       #Excel文件
    worksheet = ""      #Excel文件中的表格
    worksheetObj = None

    worksheet_col_cnt = 0
    worksheet_row_cnt = 0

    def __init__(self, workbook, worksheet):
        self.workbook = workbook
        self.worksheet = worksheet

        self.isWorkbookExist()
        self.isWorksheetExist()

        if self.worksheet in self.workbook.sheetnames:
            self.worksheetObj = self.workbook[self.worksheet]
            self.worksheet_col_cnt = self.worksheetObj.max_column
            self.worksheet_row_cnt = self.worksheetObj.max_row
        else:
            print(worksheet, " , worksheet does not exist!")
            exit()

    def isWorkbookExist(self):
        if(exists(self.workbook)):
            pass    #workbook exist
        else:
            print(self.workbook, " , file does not exist!")  #workbook not exist
            exit()

    def isWorksheetExist(self):
        try:
            self.workbook = load_workbook(self.workbook)     #check worksheet exist
        except:
            print(self.worksheet," , file not support!")     #worksheet note exist
            exit();

    def readLineWithRange(self, row, col_start, col_end):
        result = []
        for i in range(col_start, col_end + 1):
            cellObj = self.worksheetObj.cell(row = row, column = i)
            result.append(str(cellObj.value))
        return result

    def NumletterColConverter(self, col):
        col_letter = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            col_letter = chr(65 + remainder) + col_letter
        return col_letter

    def readNLine(self, lineTotal):
        result = []
        col_start = 1
        col_end = self.worksheet_col_cnt
        for i in range(1, lineTotal + 1):
            for j in range(col_start, col_end + 1):
                temp = [[],[],[]]
                temp[0] = i
                temp[1] = self.NumletterColConverter(j)
                cellObj = self.worksheetObj.cell(row = i, column = j)
                temp[2] = str(cellObj.value)
                result.append(temp)            
        return result

    def readDefaultLine(self):
        if(self.worksheet_row_cnt > 1000):
            lineTotal = 1000
        else:
            lineTotal = self.worksheet_row_cnt

        result = []
        col_start = 1
        col_end = self.worksheet_col_cnt
        for i in range(1, lineTotal + 1):
            for j in range(col_start, col_end + 1):
                temp = [[],[],[]]
                temp[0] = i
                temp[1] = self.NumletterColConverter(j)
                cellObj = self.worksheetObj.cell(row = i, column = j)
                temp[2] = str(cellObj.value)
                result.append(temp)           
        return result

    def saveData(self, result):
        tempFile = open("temp.dat","wt")
        temp = ""
        for i in range(0, len(result)):
            temp += "["
            temp += str(result[i][0])
            temp += "\\"
            temp += str(result[i][1])
            temp += "\\"
            temp += str(result[i][2])
            temp += "]"
            temp +="\n"
        tempFile.write(temp)
        tempFile.close()

def main():
    testCase3()

def testCase1():
    wb = excel(workbook="test1.xlsx", worksheet="Sheet1")
    print(wb.NumletterColConverter(12))

def testCase2():
    wb = excel(workbook="test1.xlsx", worksheet="Sheet1")
    print(wb.readDefaultLine())

def testCase3():
    wb = excel(workbook="test1.xlsx", worksheet="Sheet1")
    result = wb.readDefaultLine()
    wb.saveData(result)

if __name__ == "__main__":
    main()
