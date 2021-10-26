#############################################################################################
#Author:    Shutian Wang
#Date:      Sep. 25nd. 2021
#Desc:      Excel class
#Project:   Load Excel To Database
#############################################################################################

from typing import overload
from dns.rdatatype import NULL
from os.path import exists
from openpyxl import load_workbook

class excel():

    workbook = ""       #Excel文件
    worksheet = ""      #Excel文件中的表格
    worksheetObj = NULL

    worksheet_col_cnt = 0
    worksheet_row_cnt = 0

    def __init__(self, workbook, worksheet):
        self.workbook = workbook
        self.worksheet = worksheet

        self.isWorkbookExist()
        self.isWorksheetExist()

        if self.worksheet in self.workbook.sheetnames:
            self.worksheetObj = self.workbook[self.worksheet]
            self.worksheet_col_cnt = self.worksheetObj.max_column
            self.worksheet_row_cnt = self.worksheetObj.max_row
        else:
            print(worksheet, " , worksheet does not exist!")
            exit()

    def isWorkbookExist(self):
        if(exists(self.workbook)):
            pass    #workbook exist
        else:
            print(self.workbook, " , file does not exist!")  #workbook not exist
            exit()

    def isWorksheetExist(self):
        try:
            self.workbook = load_workbook(self.workbook)     #check worksheet exist
        except:
            print(self.worksheet," , file not support!")     #worksheet note exist
            exit();

    def letterNumRangeConverter(self, row_start, col_start, row_end, col_end):
        # summary: Use this function to convert user inputed excel worksheet start end position to an openpyxl understandable start end position
        # return [[row_start, col_start],[row_end, col_end]]
        col_start_int = 0
        col_end_int = 0
        for i in col_start:
            col_start_int = col_start_int * 26 + (ord(i.upper()) - ord('A')) + 1

        for i in col_end:
            col_end_int = col_end_int * 26 + (ord(i.upper()) - ord('A')) + 1
        
        return [[row_start, col_start_int], [row_end, col_end_int]]

    def letterNumColConverter(self, col):
        col_int = 0
        for i in col:
            col_int = col_int * 26 + (ord(i.upper()) - ord('A')) + 1

    def getCellData(self, row, col):
        # summary: return the designated cell value, for test propuse only
        cellObj = self.worksheetObj.cell(row = row, column = col)
        return cellObj.value

    def readLine(self, row):
        result = []
        for i in range(1, self.worksheet_col_cnt + 1):
            cellObj = self.worksheetObj.cell(row = row, column = i)
            result.append(str(cellObj.value))
        return result

    def readLineWithRange(self, row, col_start, col_range):
        result = []
        for i in range(col_start, col_range + 1):
            cellObj = self.worksheetObj.cell(row = row, column = i)
            result.append(str(cellObj.value))
        return result